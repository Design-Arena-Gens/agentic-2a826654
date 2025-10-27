from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .transform import HEADERS


def create_workbook(rows: Iterable[Dict[str, object]], metadata: Dict[str, object]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Companies"

    header_font = Font(bold=True)
    for column_index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(HEADERS, start=1):
            sheet.cell(row=row_index, column=column_index, value=row.get(header))

    # Auto width approximation.
    for column_index, header in enumerate(HEADERS, start=1):
        column_letter = get_column_letter(column_index)
        max_length = max(
            len(str(sheet.cell(row=row_index, column=column_index).value or ""))
            for row_index in range(1, sheet.max_row + 1)
        )
        sheet.column_dimensions[column_letter].width = min(max(len(header) + 2, max_length + 2), 60)

    sheet.freeze_panes = "A2"

    summary = workbook.create_sheet("Summary", 0)
    summary.append(["Generated at", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    for key, value in metadata.items():
        summary.append([key, value])

    for column in range(1, 3):
        column_letter = get_column_letter(column)
        summary.column_dimensions[column_letter].width = 30

    return workbook


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_xlsx_bytes(rows: List[Dict[str, object]], metadata: Dict[str, object]) -> bytes:
    workbook = create_workbook(rows, metadata)
    return workbook_to_bytes(workbook)
